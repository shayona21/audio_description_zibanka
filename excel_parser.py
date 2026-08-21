"""Parse Audio Description scripts from CSV or Excel files.

Both formats are normalized to the same list of dictionaries used by the
audio-generation pipeline.
"""

import csv
from datetime import datetime, time, timedelta
from pathlib import Path


# Frame rate for Indian TV content
FRAME_RATE = 25
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".xlsm"}


def timecode_to_ms(timecode):
    """Convert a timecode value to milliseconds.

    Text timecodes use ``HH:MM:SS:FF`` at 25 fps. Excel may return a time,
    datetime, timedelta, or numeric day fraction instead of the displayed
    string, so those native cell values are accepted as well.
    """

    if isinstance(timecode, datetime):
        timecode = timecode.time()

    if isinstance(timecode, time):
        total_seconds = (
            (timecode.hour * 3600)
            + (timecode.minute * 60)
            + timecode.second
            + (timecode.microsecond / 1_000_000)
        )
        return round(total_seconds * 1000)

    if isinstance(timecode, timedelta):
        return round(timecode.total_seconds() * 1000)

    # Excel stores times as fractions of a 24-hour day.
    if isinstance(timecode, (int, float)) and not isinstance(timecode, bool):
        if timecode < 0:
            raise ValueError("timecode cannot be negative")
        return round(timecode * 24 * 60 * 60 * 1000)

    parts = str(timecode).strip().split(":")
    if len(parts) != 4:
        raise ValueError(
            f"invalid timecode {timecode!r}; expected HH:MM:SS:FF"
        )

    try:
        hours, minutes, seconds, frames = (int(part) for part in parts)
    except ValueError as exc:
        raise ValueError(
            f"invalid timecode {timecode!r}; expected HH:MM:SS:FF"
        ) from exc

    if hours < 0 or not 0 <= minutes < 60 or not 0 <= seconds < 60:
        raise ValueError(f"invalid timecode {timecode!r}")
    if not 0 <= frames < FRAME_RATE:
        raise ValueError(
            f"invalid frame value in {timecode!r}; expected 0-{FRAME_RATE - 1}"
        )

    total_seconds = (hours * 3600) + (minutes * 60) + seconds
    return round((total_seconds + (frames / FRAME_RATE)) * 1000)


def detect_file_type(file_path):
    """Return ``csv`` or ``excel`` based on the file extension."""

    extension = Path(file_path).suffix.lower()
    if extension == ".csv":
        return "csv"
    if extension in {".xlsx", ".xls", ".xlsm"}:
        return "excel"

    supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
    raise ValueError(f"Unsupported file type. Please use one of: {supported}")


def _looks_like_header(row):
    if len(row) < 2:
        return False
    first = str(row[0] or "").strip().lower()
    second = str(row[1] or "").strip().lower()
    return "start" in first and "end" in second


def _normalize_rows(source_rows, source_name):
    """Convert CSV/Excel row values to the pipeline's stable row shape."""

    rows = []

    for source_row_number, row in enumerate(source_rows, start=1):
        row = list(row)

        if len(row) < 3:
            continue

        start_value, end_value = row[0], row[1]
        text = "" if row[2] is None else str(row[2]).strip()

        # Empty lines and rows without narration are intentionally ignored.
        if start_value is None or end_value is None or not text:
            continue
        if not str(start_value).strip() or not str(end_value).strip():
            continue
        if not rows and _looks_like_header(row):
            continue

        try:
            start_ms = timecode_to_ms(start_value)
            end_ms = timecode_to_ms(end_value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Invalid timecode on row {source_row_number} of {source_name}: {exc}"
            ) from exc

        if end_ms < start_ms:
            raise ValueError(
                f"End timecode is before start timecode on row "
                f"{source_row_number} of {source_name}"
            )

        rows.append({
            "row_number": len(rows) + 1,
            "start_ms": start_ms,
            "end_ms": end_ms,
            "duration_ms": end_ms - start_ms,
            "text": text,
        })

    if not rows:
        raise ValueError(
            "No valid audio-description rows were found. Expected start "
            "timecode, end timecode, and dialogue in the first three columns."
        )

    return rows


def parse_csv(file_path, progress_callback=print):
    """Read a UTF-8 CSV Audio Description script."""

    # utf-8-sig accepts ordinary UTF-8 and removes a BOM when spreadsheets add
    # one during CSV export.
    with open(file_path, "r", encoding="utf-8-sig", newline="") as file_handle:
        rows = _normalize_rows(csv.reader(file_handle), Path(file_path).name)

    progress_callback(f"Parsed {len(rows)} AD rows from CSV file {file_path}")
    return rows


def _xlsx_rows(file_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel .xlsx support requires the openpyxl package"
        ) from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        yield from worksheet.iter_rows(values_only=True)
    finally:
        workbook.close()


def _xls_rows(file_path):
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Excel .xls support requires the xlrd package") from exc

    workbook = xlrd.open_workbook(file_path, on_demand=True)
    try:
        worksheet = workbook.sheet_by_index(0)
        for row_number in range(worksheet.nrows):
            yield worksheet.row_values(row_number)
    finally:
        workbook.release_resources()


def parse_excel(file_path, progress_callback=print):
    """Read the first worksheet of an Excel Audio Description script."""

    extension = Path(file_path).suffix.lower()
    if extension in {".xlsx", ".xlsm"}:
        source_rows = _xlsx_rows(file_path)
    elif extension == ".xls":
        source_rows = _xls_rows(file_path)
    else:
        raise ValueError("Excel files must use .xlsx, .xlsm, or .xls")

    rows = _normalize_rows(source_rows, Path(file_path).name)
    progress_callback(f"Parsed {len(rows)} AD rows from Excel file {file_path}")
    return rows


def parse_file(file_path, progress_callback=print):
    """Detect and parse a supported CSV or Excel AD script."""

    file_type = detect_file_type(file_path)
    if file_type == "csv":
        return parse_csv(file_path, progress_callback=progress_callback)
    return parse_excel(file_path, progress_callback=progress_callback)


# Backwards-compatible, more descriptive alias for external callers.
parse_ad_file = parse_file


if __name__ == "__main__":
    parsed_rows = parse_file("raktanchal_s3_ep7_ad.csv")
    for parsed_row in parsed_rows[:5]:
        print(
            f"Row {parsed_row['row_number']}: {parsed_row['start_ms']}ms → "
            f"{parsed_row['end_ms']}ms | {parsed_row['text'][:40]}"
        )
